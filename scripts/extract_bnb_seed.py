"""
scripts/extract_bnb_seed.py
===========================
БНБ „Кредитна динамика“ (тримесечна) → `data/manual/bnb_loans_history.csv`.

Защо съществува: ЕЦБ Data Portal (набор BSI) държи българските кредитни салда
само от **01.2022** нататък. Същите редове обаче ги има при първоизточника —
БНБ ги публикува тримесечно от **2005Q4**. Този скрипт вади дългата памет от
суровия файл, който остава комитнат в репото (`data/manual/raw/`), за да е
проверимо всяко число point-in-time.

Изход (CSV):
    date        последният месец на тримесечието, ПЪРВО число (2005-12-01) —
                конвенцията на пайплайна за всички серии
    series      NFC | HH
    value_meur  млн. EUR (суровината е „хил. евро“, делено на 1000)

Дизайнът е „гърми, не гълтай“: редовете се намират ПО ЕТИКЕТ (не по фиксиран
индекс), а мерната единица на стойностния ред се сверява изрично. Ако БНБ
разбърка листа при следваща ревизия, скриптът спира с ясно съобщение вместо
да произведе тихо разместени числа.

Пускане (еднократно, при ревизия на историята — НЕ в седмичния ритъм):
    PYTHONUTF8=1 python scripts/extract_bnb_seed.py
"""
from __future__ import annotations

import argparse
import sys
from datetime import date, datetime
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parents[1]

DEFAULT_RAW = BASE_DIR / "data" / "manual" / "raw" / "cred_q_dyn_type_eur_bg_2026-07-25.xlsx"
DEFAULT_OUT = BASE_DIR / "data" / "manual" / "bnb_loans_history.csv"

SHEET_NAME = "PUB_Q_DYN_CRED_TYPE"
DATE_ROW = 2            # ред 2 носи датите на тримесечията
FIRST_DATA_COL = 3      # колона C е първото тримесечие (2005-12)
UNIT_COL = 2            # колона B носи мерната единица на реда
EXPECTED_UNIT = "хил. евро"
QUARTER_END_MONTHS = (3, 6, 9, 12)

# Етикет в колона А → ключ в изхода. Стойностите стоят на СЛЕДВАЩИЯ ред.
# „Финансови предприятия“ е отделна секция и НЕ влиза (дефиниционно различна
# от counterpart-ите 2240/2250 в BSI).
ROW_LABELS = {
    "NFC": "Нефинансови предприятия",
    "HH": "Домакинства и НТООД",
}

SERIES_DEFINITIONS = {
    "NFC": "„Нефинансови предприятия“ (ЕЦБ BSI counterpart 2240)",
    "HH": "„Домакинства и НТООД“ (ЕЦБ BSI counterpart 2250 — същата дефиниция)",
}


def _text(value) -> str:
    """Клетка → сравним текст (БНБ оставя висящ интервал след етикетите)."""
    return value.strip() if isinstance(value, str) else ""


def load_sheet(raw_path: Path):
    """Отваря листа с данните. Липсващ лист → ValueError, не тих празен изход."""
    try:
        import openpyxl
    except ImportError as e:  # pragma: no cover - зависимостта е в requirements
        raise RuntimeError(
            "За екстракцията трябва openpyxl (`pip install -r requirements.txt`)."
        ) from e

    wb = openpyxl.load_workbook(raw_path, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(
            f"Липсва лист '{SHEET_NAME}' в {raw_path.name}; намерени: {wb.sheetnames}"
        )
    return wb[SHEET_NAME]


def find_label_row(ws, label: str) -> int:
    """Номерът на реда с даден етикет в колона А. Търси се ПО ЕТИКЕТ.

    Нула или повече от едно съвпадение → ValueError: листът се е разместил и
    фиксираният индекс би върнал чужди числа.
    """
    hits = [
        r for r in range(1, ws.max_row + 1)
        if _text(ws.cell(row=r, column=1).value) == label
    ]
    if len(hits) != 1:
        raise ValueError(
            f"Етикетът '{label}' е намерен {len(hits)} пъти (очаква се точно 1) "
            f"в лист {SHEET_NAME}. Листът е разместен — провери суровината."
        )
    return hits[0]


def quarter_columns(ws) -> list[tuple[int, date]]:
    """[(колона, дата)] за всяко тримесечие; датата е първо число на месеца.

    Всяка дата се сверява, че пада в края на тримесечие (месец 3/6/9/12) —
    иначе шапката не е тази, която очакваме.
    """
    out: list[tuple[int, date]] = []
    for col in range(FIRST_DATA_COL, ws.max_column + 1):
        raw = ws.cell(row=DATE_ROW, column=col).value
        if raw is None:
            continue
        if not isinstance(raw, datetime):
            raise ValueError(
                f"Колона {col}, ред {DATE_ROW}: очакваше се дата, намерено {raw!r}."
            )
        if raw.month not in QUARTER_END_MONTHS:
            raise ValueError(
                f"Колона {col}: {raw.date()} не е край на тримесечие "
                f"(очакват се месеци {QUARTER_END_MONTHS})."
            )
        out.append((col, date(raw.year, raw.month, 1)))

    if not out:
        raise ValueError(f"Ред {DATE_ROW} не съдържа нито едно тримесечие.")
    dates = [d for _, d in out]
    if dates != sorted(dates) or len(set(dates)) != len(dates):
        raise ValueError("Тримесечията в шапката не са възходящи и уникални.")
    return out


def extract_series(ws, label: str) -> dict[date, float]:
    """Един етикет → {дата: млн. EUR}.

    Стойностите са на реда ПОД етикета; колона B на този ред трябва да казва
    „хил. евро“. Разминаване → ValueError (мерната единица е тихият убиец).
    """
    label_row = find_label_row(ws, label)
    value_row = label_row + 1
    unit = _text(ws.cell(row=value_row, column=UNIT_COL).value)
    if unit != EXPECTED_UNIT:
        raise ValueError(
            f"'{label}' (ред {label_row}): стойностният ред {value_row} казва "
            f"'{unit}' вместо '{EXPECTED_UNIT}'. Структурата на листа се е сменила."
        )

    out: dict[date, float] = {}
    for col, when in quarter_columns(ws):
        raw = ws.cell(row=value_row, column=col).value
        if isinstance(raw, bool) or not isinstance(raw, (int, float)):
            raise ValueError(
                f"'{label}' {when}: очакваше се число, намерено {raw!r} "
                f"(ред {value_row}, колона {col})."
            )
        out[when] = float(raw) / 1000.0     # хил. EUR → млн. EUR
    return out


def extract_seed(raw_path: Path = DEFAULT_RAW) -> dict[str, dict[date, float]]:
    """Суровият БНБ файл → {"NFC": {...}, "HH": {...}} в млн. EUR."""
    raw_path = Path(raw_path)
    if not raw_path.exists():
        raise FileNotFoundError(f"Липсва суровината: {raw_path}")
    ws = load_sheet(raw_path)
    return {key: extract_series(ws, label) for key, label in ROW_LABELS.items()}


def build_csv_text(
    seed: dict[str, dict[date, float]],
    raw_path: Path,
    extracted_on: date | None = None,
) -> str:
    """Готовият CSV текст — с шапка-коментари, която казва откъде идва всичко."""
    extracted_on = extracted_on or date.today()
    rel_raw = f"data/manual/raw/{Path(raw_path).name}"

    lines = [
        "# БНБ „Кредитна динамика“ — тримесечни кредитни салда (seed за дългата памет)",
        "# Източник: Българска народна банка, публикация „Кредитна динамика“,",
        f"#   лист {SHEET_NAME} (кредити по вид, „хил. евро“).",
        f"# Суров файл: {rel_raw}",
        "#   доставен от Цветослав на 25.07.2026; оригиналът остава комитнат в репото,",
        "#   за да е проверимо всяко число point-in-time.",
        f"# Извлечено на: {extracted_on.isoformat()} със scripts/extract_bnb_seed.py",
        "# Дефиниции:",
    ]
    for key in ROW_LABELS:
        lines.append(f"#   {key} = {SERIES_DEFINITIONS[key]}")
    lines += [
        "#   „Финансови предприятия“ е ОТДЕЛНА секция в листа и НЕ влиза.",
        "# Колони:",
        "#   date       = последният месец на тримесечието, първо число (2005-12-01)",
        "#   value_meur = млн. EUR (суровината е хил. EUR, делено на 1000)",
        "# Този файл НЕ се обновява седмично — ЕЦБ BSI поема напред. Нов екстракт",
        "# се прави само при ревизия на историята от БНБ.",
        "date,series,value_meur",
    ]

    for key in sorted(seed):
        for when in sorted(seed[key]):
            lines.append(f"{when.isoformat()},{key},{seed[key][when]:.3f}")

    return "\n".join(lines) + "\n"


def main(argv=None) -> int:
    parser = argparse.ArgumentParser(
        description="БНБ „Кредитна динамика“ → data/manual/bnb_loans_history.csv"
    )
    parser.add_argument("--raw", default=str(DEFAULT_RAW), help="Суровият .xlsx файл")
    parser.add_argument("--out", default=str(DEFAULT_OUT), help="Изходният CSV")
    args = parser.parse_args(argv)

    raw_path = Path(args.raw)
    seed = extract_seed(raw_path)
    text = build_csv_text(seed, raw_path)

    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(text, encoding="utf-8")

    for key, points in seed.items():
        first, last = min(points), max(points)
        print(
            f"✅ {key}: {len(points)} тримесечия · {first} → {last} · "
            f"последно {points[last]:,.3f} млн. EUR"
        )
    print(f"✅ Записано: {out_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
